import datetime
import re

from django.http import HttpResponse
from django.utils.translation import gettext, gettext_lazy as _
from openpyxl import Workbook

from security.templatetags.security import currency, format_card_number, format_sort_code, format_resolution

payment_methods = {
    'bank_transfer': _('Bank transfer'),
    'online': _('Debit card'),
}


class CreditXlsxResponse(HttpResponse):
    def __init__(self, object_list, attachment_name='export.xlsx', **kwargs):
        kwargs.setdefault(
            'content_type',
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        super().__init__(**kwargs)
        self['Content-Disposition'] = 'attachment; filename="%s"' % attachment_name
        wb = Workbook()
        write_header(wb)
        write_credits(wb, object_list)
        wb.save(self)


def write_header(workbook):
    ws = workbook.active
    headers = [
        gettext('Prisoner name'), gettext('Prisoner number'), gettext('Prison'),
        gettext('Sender name'), gettext('Payment method'),
        gettext('Bank transfer sort code'), gettext('Bank transfer account'), gettext('Bank transfer roll number'),
        gettext('Debit card number'), gettext('Debit card expiry'), gettext('Address'),
        gettext('Amount'), gettext('Date received'),
        gettext('Credited status'), gettext('Date credited'), gettext('NOMIS ID'),
        gettext('IP'),
    ]
    for col, header in enumerate(headers, start=1):
        ws.cell(column=col, row=1, value=header)


def write_credits(workbook, object_list):
    ws = workbook.active
    for row, credit in enumerate(object_list, start=2):
        cells = [
            credit['prisoner_name'],
            credit['prisoner_number'],
            credit['prison_name'],
            credit['sender_name'],
            str(payment_methods.get(credit['source'], credit['source'])),
            format_sort_code(credit['sender_sort_code']) if credit['sender_sort_code'] else '',
            credit['sender_account_number'],
            credit['sender_roll_number'],
            format_card_number(credit['card_number_last_digits']) if credit['card_number_last_digits'] else '',
            credit['card_expiry_date'],
            address_for_export(credit['billing_address']),
            currency(credit['amount']),
            credit['received_at'],
            format_resolution(credit['resolution']),
            credit['credited_at'],
            credit['nomis_transaction_id'],
            credit['ip_address'],
        ]
        for col, cell in enumerate(list(map(escape_formulae, cells)), start=1):
            ws.cell(column=col, row=row, value=cell)


def escape_formulae(value):
    """
    Escapes formulae (strings that start with =) to prevent
    spreadsheet software vulnerabilities being exploited
    :param value: the value being added to a CSV cell
    """
    if isinstance(value, str) and value.startswith('='):
        return "'" + value
    if isinstance(value, datetime.datetime):
        return value.strftime('%Y-%m-%d %H:%M:%S')
    if isinstance(value, datetime.date):
        return value.strftime('%Y-%m-%d')
    return value


def address_for_export(address):
    if not address:
        return ''
    whitespace = re.compile(r'\s+')
    keys = ('line1', 'line2', 'city', 'postcode', 'country')
    lines = (whitespace.sub(' ', address[key]).strip() for key in keys if address.get(key))
    return ', '.join(lines)
